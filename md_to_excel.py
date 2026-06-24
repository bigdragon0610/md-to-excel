#!/usr/bin/env python3
"""
md_to_excel.py — Markdown(画像付き) を「Excel上のスライド」風 .xlsx に変換する。

設計方針:
  - `---`(水平線) でスライドを区切り、1枚のシートに縦スクロールで並べる
  - 各スライドはタイトルバー + 本文ブロック(段落/箇条書き/表/画像/引用) で構成
  - スライド間にオレンジの区切り帯を入れて視覚的に分離する

依存: openpyxl, Pillow(画像サイズ取得に使用)
    pip install openpyxl pillow

使い方:
    python3 md_to_excel.py aws-news-2026-06.md aws-news-2026-06.xlsx
"""

import re
import sys
import os
import tempfile
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

try:
    from PIL import Image as PILImage, ImageDraw, ImageFont
    HAS_PIL = True
except ImportError:
    HAS_PIL = False

# ヒラギノ（macOS）→ Arial Unicode（Windows/Linux）の順で探す
_FONT_CANDIDATES = [
    "/System/Library/Fonts/ヒラギノ角ゴシック W3.ttc",
    "/System/Library/Fonts/ヒラギノ角ゴ ProN W3.ttc",
    "/System/Library/Fonts/Arial Unicode.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
]

def _pil_font(size):
    for path in _FONT_CANDIDATES:
        if os.path.exists(path):
            return ImageFont.truetype(path, size)
    return ImageFont.load_default()


# ---------- スタイル定義(スライドのテーマ) ----------
THEME = {
    "title":        "000000",  # タイトル文字色
    "text":         "000000",  # 本文文字色
    "muted":        "555555",  # 引用など補助テキスト
    "table_stripe": "F2F2F2",  # 表の縞模様（薄いグレー）
    "border":       "999999",  # 罫線
}

CANVAS_COLS      = 12
COL_WIDTH        = 11.0       # 各列の幅(文字単位)
TITLE_ROW_HEIGHT = 36
IMG_MAX_WIDTH_PX = 700


def thin_border(color=None):
    c = color or THEME["border"]
    side = Side(style="thin", color=c)
    return Border(left=side, right=side, top=side, bottom=side)


def bottom_border(color=None):
    c = color or THEME["border"]
    side = Side(style="thin", color=c)
    return Border(bottom=side)


# ---------- Markdown パーサ ----------
def split_slides(md_text):
    """Markdown を スライド単位のブロックリストに分割する。

    分割ルール: `---`(水平線) を区切りとする。
    各スライドは行のリストとして返す。
    """
    lines = md_text.splitlines()
    slides = []
    current = []
    for line in lines:
        if re.match(r"^\s*---\s*$", line):
            slides.append(current)
            current = []
        else:
            current.append(line)
    if current:
        slides.append(current)
    # 空スライドを除去
    return [s for s in slides if any(l.strip() for l in s)]


def parse_blocks(lines):
    """スライド内の行を ブロック(種類, 内容) のリストに変換する。

    ブロック種類: title / subtitle / h3 / bullet / number / quote / table / image / text / blank
    """
    blocks = []
    i = 0
    n = len(lines)
    while i < n:
        line = lines[i]
        stripped = line.strip()

        if not stripped:
            i += 1
            continue

        # 画像  ![alt](path)
        m = re.match(r"^!\[(.*?)\]\((.+?)\)\s*$", stripped)
        if m:
            blocks.append(("image", {"alt": m.group(1), "path": m.group(2)}))
            i += 1
            continue

        # 見出し
        if stripped.startswith("# "):
            blocks.append(("title", inline_text(stripped[2:])))
            i += 1
            continue
        if stripped.startswith("## "):
            blocks.append(("subtitle", inline_text(stripped[3:])))
            i += 1
            continue
        if stripped.startswith("### "):
            blocks.append(("h3", inline_text(stripped[4:])))
            i += 1
            continue

        # 表 (| ... | ... |)
        if stripped.startswith("|") and i + 1 < n and re.match(r"^\s*\|?[\s:\-|]+\|?\s*$", lines[i + 1]):
            table_lines = []
            while i < n and lines[i].strip().startswith("|"):
                table_lines.append(lines[i].strip())
                i += 1
            blocks.append(("table", parse_table(table_lines)))
            continue

        # 引用
        if stripped.startswith(">"):
            quote = inline_text(stripped.lstrip(">").strip())
            blocks.append(("quote", quote))
            i += 1
            continue

        # 箇条書き
        m = re.match(r"^(\s*)[-*]\s+(.*)$", line)
        if m:
            indent = len(m.group(1)) // 2
            blocks.append(("bullet", {"indent": indent, "text": inline_text(m.group(2))}))
            i += 1
            continue

        # 番号付きリスト
        m = re.match(r"^(\s*)\d+\.\s+(.*)$", line)
        if m:
            indent = len(m.group(1)) // 2
            blocks.append(("number", {"indent": indent, "text": inline_text(m.group(2))}))
            i += 1
            continue

        # 通常段落
        blocks.append(("text", inline_text(stripped)))
        i += 1

    return blocks


def parse_table(table_lines):
    """ '| a | b |' 形式の行リストを 2次元配列にする(区切り行は除外)。"""
    rows = []
    for idx, tl in enumerate(table_lines):
        if idx == 1:  # 区切り行 ( ---|--- ) はスキップ
            continue
        cells = [c.strip() for c in tl.strip().strip("|").split("|")]
        rows.append([inline_text(c) for c in cells])
    return rows


def inline_text(s):
    """ インライン記法を簡易処理。**bold** / *italic* / `code` / [text](url) のマーカーを除去。 """
    s = re.sub(r"\*\*(.+?)\*\*", r"\1", s)
    s = re.sub(r"\*(.+?)\*", r"\1", s)
    s = re.sub(r"`(.+?)`", r"\1", s)
    s = re.sub(r"\[(.+?)\]\((.+?)\)", r"\1", s)   # リンクはテキストだけ残す
    return s


# ---------- Excel レンダラ ----------
def render_slide(ws, blocks, base_dir, row):
    """ブロックリストを ws の row 行目から描画し、次の開始行を返す。"""
    for kind, content in blocks:
        if kind in ("title", "subtitle"):
            row = render_title(ws, row, content, is_main=(kind == "title"))
        elif kind == "h3":
            row = render_h3(ws, row, content)
        elif kind == "bullet":
            row = render_bullet(ws, row, content, marker="●")
        elif kind == "number":
            row = render_bullet(ws, row, content, marker="▸")
        elif kind == "quote":
            row = render_quote(ws, row, content)
        elif kind == "table":
            row = render_table(ws, row, content)
        elif kind == "image":
            row = render_image(ws, row, content, base_dir)
        elif kind == "text":
            row = render_text(ws, row, content)
    return row


def render_separator(ws, row):
    """スライド間に余白を挿入する。"""
    ws.row_dimensions[row].height = 24
    return row + 1


def _merge_row(ws, row, ncols=CANVAS_COLS):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    return ws.cell(row=row, column=1)


def render_title(ws, row, text, is_main):
    cell = _merge_row(ws, row)
    cell.value = text
    size = 18 if is_main else 14
    cell.font = Font(name="Meiryo", size=size, bold=True, color=THEME["title"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = TITLE_ROW_HEIGHT if is_main else TITLE_ROW_HEIGHT - 8
    return row + 2


def render_h3(ws, row, text):
    cell = _merge_row(ws, row)
    cell.value = text
    cell.font = Font(name="Meiryo", size=11, bold=True, color=THEME["text"])
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    return row + 1


def render_bullet(ws, row, content, marker="・"):
    indent = content["indent"]
    col = 1 + indent
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=CANVAS_COLS)
    cell = ws.cell(row=row, column=col)
    cell.value = ("　" * indent) + marker + content["text"]
    cell.font = Font(name="Meiryo", size=11, color=THEME["text"])
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22
    return row + 1


def render_text(ws, row, text):
    cell = _merge_row(ws, row)
    cell.value = text
    cell.font = Font(name="Meiryo", size=11, color=THEME["text"])
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22
    return row + 1


def render_quote(ws, row, text):
    cell = _merge_row(ws, row)
    cell.value = text
    cell.font = Font(name="Meiryo", size=11, color=THEME["muted"])
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=2)
    ws.row_dimensions[row].height = 24
    return row + 2


def _table_to_image(rows):
    """表を PIL Image として描画して返す。"""
    FONT_SIZE  = 14
    PAD        = 10
    ROW_HEIGHT = 36
    MIN_COL_W  = 60
    MAX_IMG_W  = 900
    BORDER_CLR = (153, 153, 153)
    HEAD_BG    = (242, 242, 242)
    WHITE      = (255, 255, 255)
    TEXT_CLR   = (0, 0, 0)

    font      = _pil_font(FONT_SIZE)
    font_bold = _pil_font(FONT_SIZE)
    ncols     = max(len(r) for r in rows)

    dummy = ImageDraw.Draw(PILImage.new("RGB", (1, 1)))
    col_widths = []
    for ci in range(ncols):
        w = MIN_COL_W
        for ri, r in enumerate(rows):
            text = r[ci] if ci < len(r) else ""
            f = font_bold if ri == 0 else font
            bbox = dummy.textbbox((0, 0), text, font=f)
            w = max(w, bbox[2] - bbox[0] + PAD * 2)
        col_widths.append(w)

    total_w = sum(col_widths)
    if total_w > MAX_IMG_W:
        scale = MAX_IMG_W / total_w
        col_widths = [max(MIN_COL_W, int(w * scale)) for w in col_widths]
        total_w = sum(col_widths)

    img = PILImage.new("RGB", (total_w + 1, ROW_HEIGHT * len(rows) + 1), WHITE)
    draw = ImageDraw.Draw(img)

    y = 0
    for ri, r in enumerate(rows):
        x = 0
        is_head = (ri == 0)
        for ci in range(ncols):
            cw = col_widths[ci]
            draw.rectangle([x, y, x + cw, y + ROW_HEIGHT], fill=HEAD_BG if is_head else WHITE)
            draw.rectangle([x, y, x + cw, y + ROW_HEIGHT], outline=BORDER_CLR)
            text = r[ci] if ci < len(r) else ""
            f = font_bold if is_head else font
            draw.text((x + PAD, y + PAD), text, font=f, fill=TEXT_CLR)
            x += cw
        y += ROW_HEIGHT

    return img


def render_table(ws, row, rows):
    if not rows:
        return row
    if not HAS_PIL:
        return row  # Pillow がない場合はスキップ

    img = _table_to_image(rows)
    tmp = tempfile.NamedTemporaryFile(suffix=".png", delete=False)
    img.save(tmp.name)
    tmp.close()

    xl_img = XLImage(tmp.name)
    xl_img.width, xl_img.height = img.width, img.height
    ws.add_image(xl_img, f"B{row}")
    rows_needed = int(img.height / 19) + 1
    for r in range(row, row + rows_needed):
        ws.row_dimensions[r].height = 19
    return row + rows_needed + 1


def render_image(ws, row, content, base_dir):
    path = os.path.join(base_dir, content["path"])
    if not os.path.exists(path):
        return render_text(ws, row, f"[画像が見つかりません: {content['path']}]")

    img = XLImage(path)
    # 表示サイズを最大幅に合わせて縮小
    if HAS_PIL:
        with PILImage.open(path) as pim:
            w, h = pim.size
    else:
        w, h = img.width, img.height
    if w > IMG_MAX_WIDTH_PX:
        scale = IMG_MAX_WIDTH_PX / w
        img.width = int(w * scale)
        img.height = int(h * scale)
    else:
        img.width, img.height = w, h

    anchor = f"{get_column_letter(2)}{row}"  # B列にアンカー(少し左マージン)
    ws.add_image(img, anchor)
    # 画像の高さ分の行を確保 (1行 ≈ 20px 換算)
    rows_needed = max(1, int(img.height / 19) + 1)
    for r in range(row, row + rows_needed):
        ws.row_dimensions[r].height = 19
    return row + rows_needed + 1


def convert(md_path, xlsx_path):
    with open(md_path, encoding="utf-8") as f:
        md_text = f.read()
    base_dir = os.path.dirname(os.path.abspath(md_path))

    slides = split_slides(md_text)
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    # 列幅を設定してキャンバス化(1シートで行う)
    for c in range(1, CANVAS_COLS + 2):
        ws.column_dimensions[get_column_letter(c)].width = COL_WIDTH
    ws.sheet_view.showGridLines = False

    row = 2  # 上に1行マージン
    for idx, slide_lines in enumerate(slides, start=1):
        blocks = parse_blocks(slide_lines)
        row = render_slide(ws, blocks, base_dir, row)
        if idx < len(slides):
            row = render_separator(ws, row)

    wb.save(xlsx_path)
    print(f"OK: {len(slides)} スライドを生成 -> {xlsx_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("usage: python3 md_to_excel.py <input.md> [output.xlsx]")
        sys.exit(1)
    md_in = sys.argv[1]
    xlsx_out = sys.argv[2] if len(sys.argv) > 2 else os.path.splitext(md_in)[0] + ".xlsx"
    convert(md_in, xlsx_out)
